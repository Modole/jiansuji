"""
数据导出相关API蓝图
"""
import os
import csv
import json
import logging
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file, current_app
from app.models.measurement import MeasurementModel
from app.models.hysteresis import HysteresisModel
from app.utils.helpers import create_response, log_api_call, now_ms, format_timestamp
import base64
from io import BytesIO

logger = logging.getLogger(__name__)

bp = Blueprint('export', __name__)


@bp.route('/api/export/csv', methods=['GET'])
def export_csv():
    """导出测量数据为CSV格式"""
    start_time = now_ms()
    
    try:
        # 获取查询参数
        data_type = request.args.get('type', 'measurements')
        start_time_param = request.args.get('start_time')
        end_time_param = request.args.get('end_time')
        limit = request.args.get('limit', 1000, type=int)
        
        # 限制最大导出数量
        limit = min(limit, 10000)
        
        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{data_type}_{timestamp}.csv"
        
        # 确保导出目录存在
        export_dir = current_app.config['EXPORT_DIR']
        os.makedirs(export_dir, exist_ok=True)
        
        filepath = os.path.join(export_dir, filename)
        
        if data_type == 'measurements':
            # 导出测量数据
            if start_time_param and end_time_param:
                data = MeasurementModel.get_measurements_by_timerange(
                    int(start_time_param), int(end_time_param), limit
                )
            else:
                data = MeasurementModel.get_latest_measurements(limit)
            
            # 写入CSV文件
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                if data:
                    fieldnames = ['timestamp', 'formatted_time'] + list(data[0].keys())
                    fieldnames = [f for f in fieldnames if f not in ['timestamp']]  # 避免重复
                    
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    
                    for row in data:
                        # 添加格式化时间
                        row_data = dict(row)
                        row_data['formatted_time'] = format_timestamp(row['timestamp'])
                        writer.writerow(row_data)
                else:
                    # 空数据时写入表头
                    writer = csv.writer(csvfile)
                    writer.writerow(['timestamp', 'formatted_time', 'message'])
                    writer.writerow([now_ms(), format_timestamp(now_ms()), 'No data available'])
        
        elif data_type == 'hysteresis':
            # 导出磁滞回线数据
            timestamps = HysteresisModel.get_hysteresis_timestamps(limit)
            
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(['timestamp', 'formatted_time', 'h_field', 'b_field', 'curve_id'])
                
                for ts in timestamps:
                    points = HysteresisModel.get_hysteresis_by_timestamp(ts)
                    formatted_time = format_timestamp(ts)
                    
                    for point in points:
                        writer.writerow([
                            ts,
                            formatted_time,
                            point.get('h_field', 0),
                            point.get('b_field', 0),
                            point.get('id', '')
                        ])
        
        else:
            error_response, status_code = create_response(
                success=False,
                error="unsupported_type",
                message=f"不支持的数据类型: {data_type}",
                status_code=400
            )
            return jsonify(error_response), status_code
        
        # 记录API调用
        duration = now_ms() - start_time
        log_api_call('/api/export/csv', 'GET', {
            'type': data_type,
            'limit': limit
        }, {'filename': filename}, duration)
        
        # 发送文件
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='text/csv'
        )
        
    except Exception as e:
        logger.error(f"导出CSV失败: {e}")
        error_response, status_code = create_response(
            success=False,
            error=str(e),
            message="导出CSV失败",
            status_code=500
        )
        return jsonify(error_response), status_code


@bp.route('/api/export/json', methods=['GET'])
def export_json():
    """导出数据为JSON格式"""
    start_time = now_ms()
    
    try:
        # 获取查询参数
        data_type = request.args.get('type', 'measurements')
        start_time_param = request.args.get('start_time')
        end_time_param = request.args.get('end_time')
        limit = request.args.get('limit', 1000, type=int)
        pretty = request.args.get('pretty', 'false').lower() == 'true'
        
        # 限制最大导出数量
        limit = min(limit, 10000)
        
        export_data = {
            'export_info': {
                'type': data_type,
                'timestamp': now_ms(),
                'formatted_time': format_timestamp(now_ms()),
                'limit': limit,
                'filters': {}
            },
            'data': []
        }
        
        # 添加过滤条件到导出信息
        if start_time_param:
            export_data['export_info']['filters']['start_time'] = int(start_time_param)
        if end_time_param:
            export_data['export_info']['filters']['end_time'] = int(end_time_param)
        
        if data_type == 'measurements':
            # 导出测量数据
            if start_time_param and end_time_param:
                data = MeasurementModel.get_measurements_by_timerange(
                    int(start_time_param), int(end_time_param), limit
                )
            else:
                data = MeasurementModel.get_latest_measurements(limit)
            
            # 添加格式化时间
            for row in data:
                row_data = dict(row)
                row_data['formatted_time'] = format_timestamp(row['timestamp'])
                export_data['data'].append(row_data)
        
        elif data_type == 'hysteresis':
            # 导出磁滞回线数据
            timestamps = HysteresisModel.get_hysteresis_timestamps(limit)
            
            for ts in timestamps:
                points = HysteresisModel.get_hysteresis_by_timestamp(ts)
                curve_data = {
                    'timestamp': ts,
                    'formatted_time': format_timestamp(ts),
                    'points': points,
                    'point_count': len(points)
                }
                export_data['data'].append(curve_data)
        
        else:
            error_response, status_code = create_response(
                success=False,
                error="unsupported_type",
                message=f"不支持的数据类型: {data_type}",
                status_code=400
            )
            return jsonify(error_response), status_code
        
        # 更新导出信息
        export_data['export_info']['record_count'] = len(export_data['data'])
        
        # 记录API调用
        duration = now_ms() - start_time
        log_api_call('/api/export/json', 'GET', {
            'type': data_type,
            'limit': limit,
            'pretty': pretty
        }, {'record_count': len(export_data['data'])}, duration)
        
        # 返回JSON数据
        if pretty:
            response = current_app.response_class(
                json.dumps(export_data, indent=2, ensure_ascii=False),
                mimetype='application/json'
            )
        else:
            response = jsonify(export_data)
        
        # 设置下载文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{data_type}_{timestamp}.json"
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        logger.error(f"导出JSON失败: {e}")
        error_response, status_code = create_response(
            success=False,
            error=str(e),
            message="导出JSON失败",
            status_code=500
        )
        return jsonify(error_response), status_code


@bp.route('/api/export/report', methods=['GET'])
def export_report():
    """导出完整测试报告"""
    start_time = now_ms()
    
    try:
        # 获取查询参数
        format_type = request.args.get('format', 'json')  # json, csv
        include_stats = request.args.get('stats', 'true').lower() == 'true'
        limit = request.args.get('limit', 1000, type=int)
        
        # 限制最大导出数量
        limit = min(limit, 5000)
        
        # 收集所有数据
        measurements = MeasurementModel.get_latest_measurements(limit)
        hysteresis_timestamps = HysteresisModel.get_hysteresis_timestamps(min(limit, 100))
        
        report_data = {
            'report_info': {
                'generated_at': now_ms(),
                'formatted_time': format_timestamp(now_ms()),
                'format': format_type,
                'include_statistics': include_stats,
                'limits': {
                    'measurements': limit,
                    'hysteresis_curves': min(limit, 100)
                }
            },
            'measurements': {
                'count': len(measurements),
                'data': measurements
            },
            'hysteresis_curves': {
                'count': len(hysteresis_timestamps),
                'data': []
            }
        }
        
        # 添加磁滞回线数据
        for ts in hysteresis_timestamps:
            points = HysteresisModel.get_hysteresis_by_timestamp(ts)
            curve_data = {
                'timestamp': ts,
                'formatted_time': format_timestamp(ts),
                'points': points,
                'point_count': len(points)
            }
            report_data['hysteresis_curves']['data'].append(curve_data)
        
        # 添加统计信息
        if include_stats:
            try:
                measurement_stats = MeasurementModel.get_measurement_stats()
                report_data['statistics'] = {
                    'measurements': measurement_stats,
                    'hysteresis': {
                        'total_curves': len(hysteresis_timestamps),
                        'total_points': sum(len(HysteresisModel.get_hysteresis_by_timestamp(ts)) 
                                          for ts in hysteresis_timestamps)
                    }
                }
            except Exception as e:
                logger.warning(f"获取统计信息失败: {e}")
                report_data['statistics'] = {'error': '统计信息获取失败'}
        
        # 记录API调用
        duration = now_ms() - start_time
        log_api_call('/api/export/report', 'GET', {
            'format': format_type,
            'include_stats': include_stats,
            'limit': limit
        }, {
            'measurements_count': len(measurements),
            'hysteresis_count': len(hysteresis_timestamps)
        }, duration)
        
        # 根据格式返回数据
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if format_type == 'csv':
            # 生成CSV报告（简化版）
            filename = f"test_report_{timestamp}.csv"
            export_dir = current_app.config['EXPORT_DIR']
            os.makedirs(export_dir, exist_ok=True)
            filepath = os.path.join(export_dir, filename)
            
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # 写入报告信息
                writer.writerow(['测试报告'])
                writer.writerow(['生成时间', format_timestamp(now_ms())])
                writer.writerow(['测量数据条数', len(measurements)])
                writer.writerow(['磁滞回线数量', len(hysteresis_timestamps)])
                writer.writerow([])
                
                # 写入测量数据
                if measurements:
                    writer.writerow(['测量数据'])
                    fieldnames = ['timestamp', 'formatted_time'] + [k for k in measurements[0].keys() if k != 'timestamp']
                    writer.writerow(fieldnames)
                    
                    for row in measurements:
                        row_data = [row['timestamp'], format_timestamp(row['timestamp'])]
                        row_data.extend([row.get(k, '') for k in fieldnames[2:]])
                        writer.writerow(row_data)
            
            return send_file(
                filepath,
                as_attachment=True,
                download_name=filename,
                mimetype='text/csv'
            )
        
        else:
            # 返回JSON报告
            filename = f"test_report_{timestamp}.json"
            response = jsonify(report_data)
            response.headers['Content-Disposition'] = f'attachment; filename={filename}'
            return response
        
    except Exception as e:
        logger.error(f"导出测试报告失败: {e}")
        error_response, status_code = create_response(
            success=False,
            error=str(e),
            message="导出测试报告失败",
            status_code=500
        )
        return jsonify(error_response), status_code


@bp.route('/api/export/static/xlsx', methods=['POST'])
def export_static_xlsx():
    """导出静态页结果为XLSX：一张数据 + 三张图表"""
    start_time = now_ms()
    try:
        payload = request.get_json(force=True) or {}
        title = payload.get('title', '静态结果导出')
        data_rows = payload.get('data_rows') or []
        charts = payload.get('charts') or []

        export_time_str = format_timestamp(now_ms())
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"静态结果_{timestamp}.xlsx"

        # 动态导入依赖，避免运行期找不到直接崩溃
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.chart import LineChart, Reference
            from openpyxl.chart.label import DataLabelList
        except Exception as e:
            error_response, status_code = create_response(
                success=False,
                error="openpyxl_not_available",
                message="服务端缺少 openpyxl/Pillow 依赖，无法生成 xlsx",
                status_code=500
            )
            return jsonify(error_response), status_code

        wb = Workbook()
        data_ws = wb.active
        data_ws.title = "数据"
        # 标题与导出时间
        data_ws.merge_cells('A1:C1')
        data_ws['A1'] = title
        data_ws['A1'].font = Font(size=14, bold=True)
        data_ws['A1'].alignment = Alignment(horizontal='center')
        data_ws['A2'] = f"导出时间: {export_time_str}"

        # 写入数据行（支持 [label, value] 数组或字典）
        row_idx = 4
        if isinstance(data_rows, list):
            for row in data_rows:
                if isinstance(row, (list, tuple)):
                    # 允许多列，常见为2列
                    for i, val in enumerate(row, start=1):
                        data_ws.cell(row=row_idx, column=i, value=str(val))
                    row_idx += 1
                elif isinstance(row, dict):
                    # 将字典按 key, value 写入两列
                    for k, v in row.items():
                        data_ws.cell(row=row_idx, column=1, value=str(k))
                        data_ws.cell(row=row_idx, column=2, value=str(v))
                        row_idx += 1
        else:
            data_ws.cell(row=row_idx, column=1, value="无数据")

        # 列宽
        for col in range(1, 6):
            data_ws.column_dimensions[get_column_letter(col)].width = 18

        # 指标图工作表（嵌入三张PNG图片）
        charts_ws = wb.create_sheet(title='指标图')
        charts_ws.merge_cells('A1:D1')
        charts_ws['A1'] = '静态指标图'
        charts_ws['A1'].font = Font(size=12, bold=True)
        charts_ws['A1'].alignment = Alignment(horizontal='center')
        charts_ws['A2'] = f"导出时间: {export_time_str}"

        def _image_from_dataurl(dataurl):
            if not (isinstance(dataurl, str) and dataurl.startswith('data:image')):
                return None
            try:
                b64 = dataurl.split(',')[1]
                img_bytes = base64.b64decode(b64)
                bio = BytesIO(img_bytes)
                return XLImage(bio)
            except Exception:
                return None

        positions = ['A4', 'I4', 'A20', 'I20']
        for i, ch in enumerate(charts[:4]):
            img = _image_from_dataurl(ch.get('image_png'))
            name = str(ch.get('name', f'图{i+1}'))
            # 图标题放置
            charts_ws.cell(row=3 if i < 2 else 19, column=(1 if i % 2 == 0 else 9), value=name)
            charts_ws.cell(row=3 if i < 2 else 19, column=(1 if i % 2 == 0 else 9)).font = Font(bold=True)
            if img:
                charts_ws.add_image(img, positions[i])

        # 合并三个趋势数据到一个工作表，按横向序号排列表格，图表一排
        summary_ws = wb.create_sheet(title='趋势数据')
        summary_ws.merge_cells('A1:D1')
        summary_ws['A1'] = f"{title} - 趋势数据汇总"
        summary_ws['A1'].font = Font(size=12, bold=True)
        summary_ws['A1'].alignment = Alignment(horizontal='center')
        summary_ws['A2'] = f"导出时间: {export_time_str}"
        summary_ws.column_dimensions['A'].width = 22

        # 收集三组数据：正向、反向、空程
        names_and_values = []
        for ch in charts:
            names_and_values.append((str(ch.get('name', '趋势数据')), ch.get('values') or []))

        def _find_series(candidates, fallback=None):
            for key in candidates:
                for nm, vals in names_and_values:
                    if key in nm:
                        return vals
            return fallback

        forward_vals = _find_series(['单向传动误差', 'unidirectional_error'])
        # 没有明确的“反向”趋势时，回退使用正向数据（仅为版式演示）
        reverse_vals = _find_series(['单向传动（反向旋转）', 'unidirectional_error_reverse', '背隙', 'backlash'], fallback=forward_vals)
        lost_vals = _find_series(['空程', 'lost_motion'])

        groups = [
            ('单向传动（正向旋转）', forward_vals, '误差'),
            ('单向传动（反向旋转）', reverse_vals, '误差'),
            ('空程', lost_vals, '空程')
        ]

        max_n = max([len(g[1]) for g in groups if g[1]] or [0])
        table_start_row = 4
        if max_n == 0:
            summary_ws.cell(row=table_start_row, column=1, value='无趋势数据')
        else:
            # 每组单独成块并列：序号/转角φ入/转角φ出-理论/实际/误差
            from openpyxl.utils import get_column_letter
            gap_cols = 2
            block_width = max_n + 1  # 1列标签 + n列数据

            def _arcmin_to_deg(x):
                try:
                    return float(x) / 60.0
                except Exception:
                    return None

            step = 360.0 / (max_n - 1) if max_n > 1 else 0.0

            for gi, (gname, gvals, err_label) in enumerate(groups[:3]):
                label_col = 1 + gi * (block_width + gap_cols)
                label_col_letter = get_column_letter(label_col)
                end_col_letter = get_column_letter(label_col + block_width)
                # 分组标题并合并单元格
                summary_ws.merge_cells(start_row=table_start_row - 1, start_column=label_col, end_row=table_start_row - 1, end_column=label_col + block_width)
                summary_ws.cell(row=table_start_row - 1, column=label_col, value=gname).font = Font(bold=True)
                summary_ws.cell(row=table_start_row - 1, column=label_col).alignment = Alignment(horizontal='center', vertical='center')

                # 行1：序号（红色）
                summary_ws.cell(row=table_start_row, column=label_col, value='序号')
                for i in range(max_n):
                    c = summary_ws.cell(row=table_start_row, column=label_col + 1 + i, value=i + 1)
                    c.font = Font(color='FF0000')
                    c.alignment = Alignment(horizontal='center', vertical='center')

                # 行2：转角φ入（°）
                summary_ws.cell(row=table_start_row + 1, column=label_col, value='转角φ入（°）')
                for i in range(max_n):
                    summary_ws.cell(row=table_start_row + 1, column=label_col + 1 + i, value=round(i * step, 0))

                # 行3：转角φ出（°）- 理论值
                summary_ws.cell(row=table_start_row + 2, column=label_col, value='转角φ出（°）- 理论值')
                for i in range(max_n):
                    summary_ws.cell(row=table_start_row + 2, column=label_col + 1 + i, value=round(i * step, 0))

                # 行4：转角φ出（°）- 实际值（理论 + 误差/空程(弧分→度)）
                summary_ws.cell(row=table_start_row + 3, column=label_col, value='转角φ出（°）- 实际值')
                for i in range(max_n):
                    theory = round(i * step, 6)
                    delta_deg = _arcmin_to_deg(gvals[i]) if gvals and i < len(gvals) else None
                    summary_ws.cell(row=table_start_row + 3, column=label_col + 1 + i, value=(theory + delta_deg) if (delta_deg is not None) else None)

                # 行5：误差/空程（原始值）
                summary_ws.cell(row=table_start_row + 4, column=label_col, value='误差/空程')
                for i in range(max_n):
                    v = gvals[i] if gvals and i < len(gvals) else None
                    summary_ws.cell(row=table_start_row + 4, column=label_col + 1 + i, value=v)

                # 在该分组下方插入图表（取误差/空程行）
                try:
                    chart = LineChart()
                    chart.title = gname
                    chart.style = 13
                    chart.y_axis.title = '值'
                    chart.x_axis.title = '样本序号'
                    data_ref = Reference(summary_ws, min_col=label_col + 1, min_row=table_start_row + 4, max_col=label_col + max_n, max_row=table_start_row + 4)
                    cat_ref = Reference(summary_ws, min_col=label_col + 1, min_row=table_start_row, max_col=label_col + max_n, max_row=table_start_row)
                    chart.add_data(data_ref, from_rows=True)
                    chart.set_categories(cat_ref)
                    if chart.series:
                        chart.series[0].data_labels = DataLabelList()
                        chart.series[0].data_labels.showVal = True
                    chart.height = 12
                    chart.width = 16  # 缩窄避免重叠
                    anchor = f"{label_col_letter}{table_start_row + 6}"
                    summary_ws.add_chart(chart, anchor)
                except Exception:
                    pass

        # 写到内存并返回
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        duration = now_ms() - start_time
        log_api_call('/api/export/static/xlsx', 'POST', {
            'charts': len(charts)
        }, {
            'filename': filename
        }, duration)

        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"导出静态结果XLSX失败: {e}")
        error_response, status_code = create_response(
            success=False,
            error=str(e),
            message="导出静态结果XLSX失败",
            status_code=500
        )
        return jsonify(error_response), status_code


@bp.route('/api/export/hysteresis/xlsx', methods=['POST'])
def export_hysteresis_xlsx():
    """导出滞回曲线为带嵌入图片的XLSX"""
    start_time = now_ms()
    try:
        payload = request.get_json(force=True) or {}
        title = payload.get('title', '滞回曲线导出')
        datasets = payload.get('datasets', [])
        image_dataurl = payload.get('image_png')

        # 数据来源选择：优先使用前端传来的“本次记录”数据；当为空或显式要求使用数据库时，才读取数据库最新记录
        prefer_db = bool(payload.get('prefer_db') or payload.get('use_db'))
        def _datasets_count(ds):
            try:
                total = 0
                for d in ds or []:
                    arr = d.get('pts') or d.get('rows') or []
                    total += len(arr)
                return total
            except Exception:
                return 0
        try:
            if prefer_db or _datasets_count(datasets) == 0:
                from app.services.data_service import DataService
                recorded_forward = DataService.get_hysteresis_curve_data('forward') or []
                recorded_reverse = DataService.get_hysteresis_curve_data('reverse') or []
                recorded_full = DataService.get_hysteresis_curve_data('hysteresis') or []
                # 如果数据库有记录，则覆盖前端传入的 datasets
                if recorded_forward or recorded_reverse or recorded_full:
                    def _chunk(arr, size):
                        return [arr[i:i+size] for i in range(0, len(arr), size)]
                    page_size = 1000
                    datasets = []
                    if recorded_full:
                        for i, c in enumerate(_chunk(recorded_full, page_size)):
                            datasets.append({ 'name': f'完整-记录-页{i+1}', 'pts': c })
                    if recorded_forward:
                        for i, c in enumerate(_chunk(recorded_forward, page_size)):
                            datasets.append({ 'name': f'正向-记录-页{i+1}', 'pts': c })
                    if recorded_reverse:
                        for i, c in enumerate(_chunk(recorded_reverse, page_size)):
                            datasets.append({ 'name': f'反向-记录-页{i+1}', 'pts': c })
        except Exception as e:
            logger.warning(f"读取记录数据失败，回退使用前端数据: {e}")
        export_time_str = format_timestamp(now_ms())
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"滞回曲线_{timestamp}.xlsx"

        # 动态导入依赖，避免运行期找不到直接崩溃
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.chart import LineChart, Reference
            from openpyxl.chart.series import SeriesLabel
            from openpyxl.chart.label import DataLabelList
        except Exception as e:
            logger.error(f"openpyxl/Pillow 不可用: {e}")
            error_response, status_code = create_response(
                success=False,
                error="openpyxl_not_available",
                message="服务端缺少 openpyxl/Pillow 依赖，无法生成带图片的 xlsx",
                status_code=500
            )
            return jsonify(error_response), status_code

        wb = Workbook()
        summary_ws = wb.active
        summary_ws.title = "总览"
        # 标题
        summary_ws.merge_cells('A1:D1')
        summary_ws['A1'] = title
        summary_ws['A1'].font = Font(size=14, bold=True)
        summary_ws['A1'].alignment = Alignment(horizontal='center')
        summary_ws['A2'] = f"导出时间: {export_time_str}"

        # 嵌入图表图片（若提供）
        if isinstance(image_dataurl, str) and image_dataurl.startswith('data:image'):
            try:
                b64 = image_dataurl.split(',')[1]
                img_bytes = base64.b64decode(b64)
                bio = BytesIO(img_bytes)
                xl_img = XLImage(bio)
                # 放置在 A4 位置
                summary_ws.add_image(xl_img, 'A4')
            except Exception as e:
                logger.warning(f"插入图像失败: {e}")

        # 新增工作表：转角表（上/中/下三组转角数据）
        try:
            # 规范化点结构为 {'angle': float, 'torque': float}
            def _norm_pts(pts):
                out = []
                for p in pts or []:
                    if isinstance(p, dict):
                        angle = p.get('angle') if p.get('angle') is not None else p.get('x', 0)
                        torque = p.get('torque') if p.get('torque') is not None else p.get('y', 0)
                    elif isinstance(p, (list, tuple)) and len(p) >= 2:
                        angle, torque = p[0], p[1]
                    else:
                        angle, torque = 0, 0
                    try:
                        out.append({'angle': float(angle), 'torque': float(torque)})
                    except Exception:
                        out.append({'angle': 0.0, 'torque': 0.0})
                return out

            forward_all, reverse_all, full_all = [], [], []
            for dsi in datasets:
                nm = str(dsi.get('name', ''))
                pts_i = dsi.get('pts') or dsi.get('rows') or []
                norm = _norm_pts(pts_i)
                if '正向' in nm:
                    forward_all.extend(norm)
                elif '反向' in nm:
                    reverse_all.extend(norm)
                else:
                    full_all.extend(norm)

            # 若前后向缺失，尝试从完整曲线拆分（根据角位移的上/下趋势）
            if (not forward_all or not reverse_all) and full_all:
                prev_angle = full_all[0]['angle']
                for p in full_all:
                    if p['angle'] >= prev_angle:
                        forward_all.append(p)
                    else:
                        reverse_all.append(p)
                    prev_angle = p['angle']

            # 线性插值角度值（按扭矩）
            def _interp_angle(points, t):
                if not points:
                    return None
                arr = sorted(points, key=lambda x: x['torque'])
                if t <= arr[0]['torque']:
                    return arr[0]['angle']
                if t >= arr[-1]['torque']:
                    return arr[-1]['angle']
                for i in range(len(arr) - 1):
                    a, b = arr[i], arr[i + 1]
                    ta, tb = a['torque'], b['torque']
                    if (ta <= t <= tb) or (tb <= t <= ta):
                        denom = (tb - ta) if (tb - ta) != 0 else 1e-9
                        r = (t - ta) / denom
                        return a['angle'] + r * (b['angle'] - a['angle'])
                return arr[-1]['angle']

            # 计算列数与扭矩序列（按记录点生成列，不再固定14）
            base_series = forward_all if len(forward_all) >= len(reverse_all) else reverse_all
            if not base_series:
                base_series = full_all
            torque_steps = sorted({p['torque'] for p in base_series})
            # 若仍为空，则退回到 14 等分
            if torque_steps:
                steps = torque_steps
                ncols = len(steps)
            else:
                ncols = 14
                all_torques = [p['torque'] for p in (forward_all + reverse_all + full_all)]
                t_min = 0.0
                t_max = max(all_torques) if all_torques else 0.0
                if t_max <= t_min:
                    t_max = t_min + 1.0
                steps = [t_min + (t_max - t_min) * i / (ncols - 1 if ncols > 1 else 1) for i in range(ncols)]

            # 计算上/下/中三组
            up_vals = [_interp_angle(forward_all, t) for t in steps]
            down_vals = [_interp_angle(reverse_all, t) for t in steps]
            mid_vals = []
            for u, d in zip(up_vals, down_vals):
                if u is None and d is None:
                    mid_vals.append(None)
                elif u is None:
                    mid_vals.append(d)
                elif d is None:
                    mid_vals.append(u)
                else:
                    mid_vals.append((u + d) / 2.0)

            # 写入工作表
            angle_ws = wb.create_sheet(title='转角表')
            # 第一行：列序号 1..n
            angle_ws.cell(row=1, column=1, value='')
            for i in range(ncols):
                c = angle_ws.cell(row=1, column=i + 2, value=i + 1)
                c.alignment = Alignment(horizontal='center', vertical='center')
            # 第二行：扭矩值
            c0 = angle_ws.cell(row=2, column=1, value='扭矩值\n(NM)')
            c0.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for i, tv in enumerate(steps):
                c = angle_ws.cell(row=2, column=i + 2, value=round(tv, 2))
                c.alignment = Alignment(horizontal='center', vertical='center')
            # 第三行：上（红色）
            c1 = angle_ws.cell(row=3, column=1, value='上\n转角φ\n(″)')
            c1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for i, v in enumerate(up_vals):
                vv = round(v, 2) if isinstance(v, (int, float)) else None
                c = angle_ws.cell(row=3, column=i + 2, value=vv)
                c.font = Font(color='FFDC2626')
                c.alignment = Alignment(horizontal='center', vertical='center')
            # 第四行：下
            c2 = angle_ws.cell(row=4, column=1, value='下\n转角φ\n(″)')
            c2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for i, v in enumerate(down_vals):
                vv = round(v, 2) if isinstance(v, (int, float)) else None
                c = angle_ws.cell(row=4, column=i + 2, value=vv)
                c.alignment = Alignment(horizontal='center', vertical='center')
            # 第五行：中
            c3 = angle_ws.cell(row=5, column=1, value='中\n转角φ\n(″)')
            c3.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for i, v in enumerate(mid_vals):
                vv = round(v, 2) if isinstance(v, (int, float)) else None
                c = angle_ws.cell(row=5, column=i + 2, value=vv)
                c.alignment = Alignment(horizontal='center', vertical='center')

            # 列宽与边框
            angle_ws.column_dimensions['A'].width = 12
            for i in range(ncols):
                angle_ws.column_dimensions[get_column_letter(i + 2)].width = 10
            thin = Side(style='thin', color='FFCBD5E1')
            for r in range(1, 6):
                for c in range(1, ncols + 2):
                    cell = angle_ws.cell(row=r, column=c)
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # 创建并插入折线图
            try:
                chart = LineChart()
                chart.title = '滞回曲线'
                chart.style = 2
                # X 轴: 扭矩
                xref = Reference(angle_ws, min_col=2, min_row=2, max_col=ncols + 1, max_row=2)
                # Y 系列: 上/中/下
                y_up = Reference(angle_ws, min_col=2, min_row=3, max_col=ncols + 1, max_row=3)
                y_mid = Reference(angle_ws, min_col=2, min_row=5, max_col=ncols + 1, max_row=5)
                y_down = Reference(angle_ws, min_col=2, min_row=4, max_col=ncols + 1, max_row=4)
                # 按行作为单个系列添加（避免按列拆成多系列）
                chart.add_data(y_up, titles_from_data=False, from_rows=True)
                chart.add_data(y_mid, titles_from_data=False, from_rows=True)
                chart.add_data(y_down, titles_from_data=False, from_rows=True)
                # 将 X 轴设为扭矩
                chart.set_categories(xref)
                # 不显示节点数值标签（原生折线图）
                chart.series[0].title = SeriesLabel(v="上")
                chart.series[1].title = SeriesLabel(v="中")
                chart.series[2].title = SeriesLabel(v="下")
                chart.x_axis.title = '扭矩 (Nm)'
                chart.y_axis.title = '转角φ (″)'
                chart.height = 14
                chart.width = 18
                angle_ws.add_chart(chart, 'B10')

                # 兼容回退：生成 PNG 并插入（WPS/部分查看器可能不显示原生图表）
                try:
                    import matplotlib
                    matplotlib.use('Agg')
                    import matplotlib.pyplot as plt
                    import numpy as np
                    # 修复中文乱码：设置常见中文字体，禁用负号乱码
                    plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'Arial Unicode MS', 'DejaVu Sans']
                    plt.rcParams['axes.unicode_minus'] = False

                    def _series_xy(xs, ys):
                        xv, yv = [], []
                        for x, y in zip(xs, ys):
                            if y is None:
                                continue
                            xv.append(x)
                            yv.append(y)
                        return xv, yv

                    fig, ax = plt.subplots(figsize=(8, 5), dpi=100)
                    x = steps
                    xu, yu = _series_xy(x, up_vals)
                    xm, ym = _series_xy(x, mid_vals)
                    xd, yd = _series_xy(x, down_vals)
                    if xu:
                        ax.plot(xu, yu, '-o', label='上', color='#dc2626', markersize=3)
                        for xx, yy in zip(xu, yu):
                            ax.annotate(f'{yy:.2f}', xy=(xx, yy), textcoords='offset points', xytext=(0, 4), ha='center', va='bottom', fontsize=8, color='#dc2626')
                    if xm:
                        ax.plot(xm, ym, '-o', label='中', color='#16a34a', markersize=3)
                        for xx, yy in zip(xm, ym):
                            ax.annotate(f'{yy:.2f}', xy=(xx, yy), textcoords='offset points', xytext=(0, 4), ha='center', va='bottom', fontsize=8, color='#16a34a')
                    if xd:
                        ax.plot(xd, yd, '-o', label='下', color='#2563eb', markersize=3)
                        for xx, yy in zip(xd, yd):
                            ax.annotate(f'{yy:.2f}', xy=(xx, yy), textcoords='offset points', xytext=(0, 4), ha='center', va='bottom', fontsize=8, color='#2563eb')
                    ax.set_title('滞回曲线')
                    ax.set_xlabel('扭矩 (Nm)')
                    ax.set_ylabel('转角φ (″)')
                    ax.grid(True, linestyle='--', alpha=0.3)
                    ax.legend()

                    buf_png = BytesIO()
                    fig.tight_layout()
                    fig.savefig(buf_png, format='png')
                    plt.close(fig)
                    buf_png.seek(0)
                    xl_img = XLImage(buf_png)
                    # 并排放置到同一行，避免与原生图表重叠
                    angle_ws.add_image(xl_img, 'N10')
                except Exception as ie:
                    logger.warning(f"插入PNG图表回退失败: {ie}")

            except Exception as ce:
                logger.warning(f"插入折线图失败: {ce}")
        except Exception as e:
            logger.warning(f"生成‘转角表’工作表失败: {e}")

        # 为每个数据集创建工作表
        for ds in datasets:
            name = str(ds.get('name', '数据'))[:31]
            pts = ds.get('pts') or ds.get('rows') or []
            ws = wb.create_sheet(title=name)
            ws.merge_cells('A1:D1')
            ws['A1'] = f"{title} - {name}"
            ws['A1'].font = Font(size=12, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws['A2'] = f"导出时间: {export_time_str}"
            ws['A3'] = "说明: 序号、角位移(度)、扭矩(N·m)"
            ws.append(["序号", "角位移", "扭矩"])  # 列头

            idx = 1
            for p in pts:
                if isinstance(p, dict):
                    angle = p.get('angle') if p.get('angle') is not None else p.get('x', 0)
                    torque = p.get('torque') if p.get('torque') is not None else p.get('y', 0)
                elif isinstance(p, (list, tuple)) and len(p) >= 2:
                    angle, torque = p[0], p[1]
                else:
                    angle, torque = 0, 0
                ws.append([idx, angle, torque])
                idx += 1

            # 列宽
            widths = [6, 12, 12, 12]
            for c in range(1, len(widths) + 1):
                ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]

        # 写到内存并返回
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        duration = now_ms() - start_time
        log_api_call('/api/export/hysteresis/xlsx', 'POST', {
            'datasets': len(datasets)
        }, {
            'filename': filename
        }, duration)

        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"导出Hysteresis XLSX失败: {e}")
        error_response, status_code = create_response(
            success=False,
            error=str(e),
            message="导出XLSX失败",
            status_code=500
        )
        return jsonify(error_response), status_code